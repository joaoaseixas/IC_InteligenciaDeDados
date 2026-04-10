import json
import re
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import ARTICLES_DIR, CHECKPOINT_FILE

XLSX_PATH = Path(ARTICLES_DIR) / "artigos.xlsx"
COLUMNS = ["Título", "Autor", "Data", "Link", "DOI", "PDF"]


def _setup_sheet(ws):
    ws.append(COLUMNS)
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 40


def save_article(article: dict) -> Path:
    Path(ARTICLES_DIR).mkdir(parents=True, exist_ok=True)
    if XLSX_PATH.exists():
        wb = load_workbook(XLSX_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Artigos"
        _setup_sheet(ws)

    ws.append([
        article.get("title", ""),
        article.get("author", ""),
        article.get("date", ""),
        article.get("url", ""),
        article.get("doi", "") or "",
        article.get("pdf_link", "") or "",
    ])
    wb.save(XLSX_PATH)
    return XLSX_PATH


def save_index(articles: list[dict]) -> None:
    print(f"[ÍNDICE] {len(articles)} artigos salvos em {XLSX_PATH}")


# --- Checkpoint ---

def save_checkpoint(page: int, seen_urls: set, collected: list) -> None:
    Path(CHECKPOINT_FILE).parent.mkdir(parents=True, exist_ok=True)
    data = {"last_page": page, "seen_urls": list(seen_urls), "collected": collected}
    with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def load_checkpoint() -> dict | None:
    path = Path(CHECKPOINT_FILE)
    if not path.exists():
        return None
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def clear_checkpoint() -> None:
    path = Path(CHECKPOINT_FILE)
    if path.exists():
        path.unlink()
