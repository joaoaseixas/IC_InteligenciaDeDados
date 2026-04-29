"""
Extrai textos dos artigos já coletados e salva em CSV.
Pode ler do XLSX gerado pelo scraper OU de arquivos enviados manualmente (xlsx/csv).

Uso:
    python extract_texts.py                        # usa artigos.xlsx padrão
    python extract_texts.py arquivo_externo.xlsx   # usa arquivo informado
"""

import sys
import csv
import unicodedata
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, ".")

from scraper.article_scraper import get_article_content
from config import ARTICLES_DIR

DEFAULT_XLSX = Path(ARTICLES_DIR) / "artigos.xlsx"
OUTPUT_CSV = Path(ARTICLES_DIR) / "textos_extraidos.csv"


def _normalize(text: str) -> str:
    return unicodedata.normalize("NFD", text).encode("ascii", "ignore").decode().lower()


def load_articles_from_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    articles = []
    for row in rows[1:]:
        item = dict(zip(headers, row))
        url = item.get("link") or item.get("url") or item.get("link do artigo") or ""
        title = item.get("título") or item.get("titulo") or item.get("title") or ""
        if url:
            articles.append({"title": str(title).strip(), "url": str(url).strip()})
    return articles


def load_articles_from_csv(path: Path) -> list[dict]:
    articles = []
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            url = row.get("link") or row.get("url") or row.get("Link") or row.get("URL") or ""
            title = row.get("título") or row.get("titulo") or row.get("Título") or row.get("title") or ""
            if url:
                articles.append({"title": title.strip(), "url": url.strip()})
    return articles


def extract_and_save(articles: list[dict]) -> None:
    Path(ARTICLES_DIR).mkdir(parents=True, exist_ok=True)
    total = len(articles)
    print(f"\n[INFO] {total} artigos para processar...\n")

    with open(OUTPUT_CSV, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["titulo", "autor", "data", "url", "doi", "texto"])
        writer.writeheader()

        for i, meta in enumerate(articles, 1):
            print(f"[{i}/{total}] {meta['title'][:80]}")
            content = get_article_content(meta)
            if not content or not content.get("text"):
                print("  [AVISO] Sem texto extraído, pulando.")
                continue
            writer.writerow({
                "titulo": content.get("title", ""),
                "autor": content.get("author", ""),
                "data": content.get("date", ""),
                "url": content.get("url", ""),
                "doi": content.get("doi", "") or "",
                "texto": content.get("text", "").replace("\n", " "),
            })

    print(f"\n[OK] CSV salvo em: {OUTPUT_CSV}")


def _load_file(path: Path) -> list[dict]:
    ext = path.suffix.lower()
    if ext == ".xlsx":
        return load_articles_from_xlsx(path)
    elif ext == ".csv":
        return load_articles_from_csv(path)
    else:
        print(f"  [ERRO] Formato não suportado: {ext}. Use .xlsx ou .csv")
        return []


def main():
    # Modo direto: python extract_texts.py arquivo.xlsx
    if len(sys.argv) > 1:
        source = Path(sys.argv[1])
        if not source.exists():
            print(f"[ERRO] Arquivo não encontrado: {source}")
            sys.exit(1)
        articles = _load_file(source)
        if not articles:
            print("[AVISO] Nenhum artigo encontrado no arquivo.")
            sys.exit(0)
        extract_and_save(articles)
        return

    # Modo interativo: adicionar múltiplos arquivos
    print("\n=== Extrator de Textos ===")
    print(f"[PADRÃO] {DEFAULT_XLSX}")
    print("\nOpções:")
    print("  [1] Usar artigos.xlsx padrão")
    print("  [2] Adicionar arquivos manualmente")

    escolha = input("\nEscolha (1 ou 2): ").strip()

    if escolha == "1":
        if not DEFAULT_XLSX.exists():
            print(f"[ERRO] Arquivo padrão não encontrado: {DEFAULT_XLSX}")
            sys.exit(1)
        articles = _load_file(DEFAULT_XLSX)

    elif escolha == "2":
        arquivos = []
        print("\nDigite o caminho completo de cada arquivo (Enter em branco para finalizar):")
        while True:
            caminho = input(f"  Arquivo {len(arquivos) + 1}: ").strip().strip('"')
            if not caminho:
                if not arquivos:
                    print("[AVISO] Nenhum arquivo adicionado.")
                    sys.exit(0)
                break
            p = Path(caminho)
            if not p.exists():
                print(f"  [ERRO] Não encontrado: {p}")
                continue
            arquivos.append(p)
            print(f"  [OK] Adicionado: {p.name}")

        print(f"\n[INFO] {len(arquivos)} arquivo(s) carregado(s).")
        articles = []
        seen_urls = set()
        for arq in arquivos:
            novos = _load_file(arq)
            antes = len(articles)
            for a in novos:
                if a["url"] not in seen_urls:
                    seen_urls.add(a["url"])
                    articles.append(a)
            print(f"  {arq.name}: {len(articles) - antes} artigos únicos adicionados")
    else:
        print("[ERRO] Opção inválida.")
        sys.exit(1)

    if not articles:
        print("[AVISO] Nenhum artigo encontrado nos arquivos.")
        sys.exit(0)

    extract_and_save(articles)


if __name__ == "__main__":
    main()
